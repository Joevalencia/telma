# -*- coding: utf-8 -*-
"""
Created on Wed Mar 20 18:23:04 2024

@author: jvalenciaf001
"""
import pandas as pd
import os
from tqdm import tqdm
import seaborn as sns
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import matplotlib.pyplot as plt
sns.set_style('darkgrid')

print('INICIALIZACIÓN DE EXPORTACION FICHEROS EN "Resultado_cierre" DEL RESUMEN DE LAS COMISIONES POR PRODUCTO')


ruta_script = os.path.dirname(__file__)
#ruta_script = r'C:\Users\jvalenciaf001\Desktop\PwC\Clienti\Generali\Progetto UL\Data\13. Cierre mensual UL\2023'
ruta_file_feather_cvul = 'data_converted_CVUL'
ruta_feather_cvul = os.path.join(ruta_script, ruta_file_feather_cvul)

## RUTA DE RESULTADOS
ruta_results = os.path.join(ruta_script, 'Resultado_cierre')
ruta_cvul_res = os.path.join(ruta_results, 'Comisiones')
if not os.path.exists(ruta_cvul_res):
    os.makedirs(ruta_cvul_res)


# File feather dentro la carpeta de datos convertidos a feather CVUL
all_file_cvul = os.listdir(ruta_feather_cvul)



# Funcion donde de guardaran en forma de dictionary las comisiones por producto

def comisiones_producto(cvul, patr):
   
    dict_resultados = {
        producto: cvul[cvul.POLIZA.isin(patr.drop_duplicates(subset='POLIZA', keep='first').query(f'PRODUCTO == "{producto}"').POLIZA.unique())].IMPORTE.sum().round(2) 
        for producto in patr.PRODUCTO.unique()
    }
    df_comm = pd.DataFrame([dict_resultados]).T.sort_values(by=[0], ascending =False)
    return df_comm



# File feather de ruta a patrimonio UL
try:
    ruta_others_feather = os.path.join(ruta_script, 'data_converted')
    ruta_patrul = os.path.join(ruta_others_feather, 'PATRIMONIO_UL.feather')
    patrimonio_ul_feather = pd.read_feather(ruta_patrul)
except Exception:
    print('NO SE HA PODIDO IMPORTAR EL ARCHIVO PATRIMUL.feather')



# Exportacion ficheros trimestrales
## Q1

try: 
    with tqdm(total=100, desc="Generando y exportando comisiones por producto CVUL_Q1") as pbar:
        
        df_cvul_03 = pd.read_feather(os.path.join(ruta_feather_cvul, all_file_cvul[1]))
        summary_com = comisiones_producto(df_cvul_03, patrimonio_ul_feather)
        
        df_summ = pd.DataFrame({'Productos': summary_com.index,
                                'COMISIONES':summary_com[0]})
        pbar.update(33)
        
        # Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value='Productos') 
        sheet.cell(row=1, column=2, value='COMISIONES')
        
        # Guardar el DataFrame en el libro de Excel
        for i, row in enumerate(df_summ.iterrows(), start=1):
            for j, value in enumerate(row[1], start=1):
                sheet.cell(row=i+1, column=j, value=value)
        pbar.update(33)
        
        chart = BarChart()
        chart.title = "Comisiones 1º trimestre por producto"
        chart.x_axis.title = "Producto"
        chart.y_axis.title = "Importes_Q1"
        
        data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=11)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=11)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        sheet.add_chart(chart, "D2")
        
        workbook.save(os.path.join(ruta_cvul_res, 'Comisiones_Q1.xlsx'))
                
            
        pbar.update(34)
        #summary_com.rename(columns={0: "Importes_Q1"})
        # new_sum = pd.DataFrame({'Importes_Q1': list(summary_com[0])})
        # pbar.update(50)
        # summary_com.to_csv(os.path.join(ruta_cvul_res,'Comisiones_Q1.csv'))
        # fig1 = sns.barplot(data=new_sum.head(10), x=summary_com.index[:10], y=new_sum.columns[0])
        # plt.title('Comisiones 1º trimestre por producto ', fontweight = 'bold')
        # plt.savefig(os.path.join(ruta_cvul_res,'G_Comisiones_Q1.jpg'))
        # plt.close()
        # pbar.update(50)

except Exception:
    print('NO SE HA PODIDO IMPORTAR EL ARCHIVO CVUL_Q1')


##Q2

try: 
    with tqdm(total=100, desc="Generando y exportando comisiones por producto CVUL_Q2") as pbar:
            
        df_cvul_06 = pd.read_feather(os.path.join(ruta_feather_cvul, all_file_cvul[3]))
        summary_com = comisiones_producto(df_cvul_06, patrimonio_ul_feather)
        
        df_summ = pd.DataFrame({'Productos': summary_com.index,
                                'COMISIONES':summary_com[0]})
        pbar.update(33)
        
        # Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value='Productos') 
        sheet.cell(row=1, column=2, value='COMISIONES')
        
        # Guardar el DataFrame en el libro de Excel
        for i, row in enumerate(df_summ.iterrows(), start=1):
            for j, value in enumerate(row[1], start=1):
                sheet.cell(row=i+1, column=j, value=value)
        pbar.update(33)
        
        chart = BarChart()
        chart.title = "Comisiones 2º trimestre por producto"
        chart.x_axis.title = "Producto"
        chart.y_axis.title = "Importes_Q2"
        
        data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=11)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=11)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        sheet.add_chart(chart, "D2")
        
        workbook.save(os.path.join(ruta_cvul_res, 'Comisiones_Q2.xlsx'))
                
            
        pbar.update(34)
        #summary_com.rename(columns={0: "Importes_Q1"})
        # new_sum = pd.DataFrame({'Importes_Q2': list(summary_com[0])})
        # pbar.update(50)
        # summary_com.to_csv(os.path.join(ruta_cvul_res,'Comisiones_Q2.csv'))
        # fig1 = sns.barplot(data=new_sum.head(10), x=summary_com.index[:10], y=new_sum.columns[0])
        # plt.title('Comisiones 2º trimestre por producto', fontweight = 'bold')
        # plt.savefig(os.path.join(ruta_cvul_res,'G_Comisiones_Q2.jpg'))
        # plt.close()
        # pbar.update(50)
    
except Exception:
    print('NO SE HA PODIDO IMPORTAR EL ARCHIVO CVUL_Q2')


##Q3

try: 
    with tqdm(total=100, desc="Generando y exportando comisiones por producto CVUL_Q3") as pbar:
        
            
        # df_cvul_09 = pd.read_feather(os.path.join(ruta_feather_cvul, all_file_cvul[5]))
        # summary_com = comisiones_producto(df_cvul_09, patrimonio_ul_feather)
        # #summary_com.rename(columns={0: "Importes_Q1"})
        # new_sum = pd.DataFrame({'Importes_Q3': list(summary_com[0])})
        # pbar.update(50)
        # summary_com.to_csv(os.path.join(ruta_cvul_res,'Comisiones_Q3.csv'))
        # fig1 = sns.barplot(data=new_sum.head(10), x=summary_com.index[:10], y=new_sum.columns[0])
        # plt.title('Comisiones 3º trimestre por producto', fontweight = 'bold')
        # plt.savefig(os.path.join(ruta_cvul_res,'G_Comisiones_Q3.jpg'))
        # plt.close()
        
        df_cvul_09 = pd.read_feather(os.path.join(ruta_feather_cvul, all_file_cvul[5]))
        pbar.update(33)
            
                
        summary_com = comisiones_producto(df_cvul_09, patrimonio_ul_feather)
        
        df_summ = pd.DataFrame({'Productos': summary_com.index,
                                'COMISIONES':summary_com[0]})
        
        # Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value='Productos') 
        sheet.cell(row=1, column=2, value='COMISIONES')
        
        # Guardar el DataFrame en el libro de Excel
        for i, row in enumerate(df_summ.iterrows(), start=1):
            for j, value in enumerate(row[1], start=1):
                sheet.cell(row=i+1, column=j, value=value)
        pbar.update(33)
        
        chart = BarChart()
        chart.title = "Comisiones 3º trimestre por producto"
        chart.x_axis.title = "Producto"
        chart.y_axis.title = "Importes_Q3"
        
        data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=11)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=11)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        sheet.add_chart(chart, "D2")
        
        workbook.save(os.path.join(ruta_cvul_res, 'Comisiones_Q3.xlsx'))
                
            
        pbar.update(34)
        
except Exception:
    print('NO SE HA PODIDO IMPORTAR EL ARCHIVO CVUL_Q3')


##Q4

try: 
    with tqdm(total=100, desc="Generando y exportando comisiones por producto CVUL_Q4") as pbar:
            
        df_cvul_12 = pd.read_feather(os.path.join(ruta_feather_cvul, all_file_cvul[7]))
        summary_com = comisiones_producto(df_cvul_12, patrimonio_ul_feather)
        
        df_summ = pd.DataFrame({'Productos': summary_com.index,
                                'COMISIONES':summary_com[0]})
        pbar.update(33)
        
        # Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value='Productos') 
        sheet.cell(row=1, column=2, value='COMISIONES')
        
        # Guardar el DataFrame en el libro de Excel
        for i, row in enumerate(df_summ.iterrows(), start=1):
            for j, value in enumerate(row[1], start=1):
                sheet.cell(row=i+1, column=j, value=value)
        pbar.update(33)
        
        chart = BarChart()
        chart.title = "Comisiones 4º trimestre por producto"
        chart.x_axis.title = "Producto"
        chart.y_axis.title = "Importes_Q4"
        
        data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=11)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=11)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        sheet.add_chart(chart, "D2")
        
        workbook.save(os.path.join(ruta_cvul_res, 'Comisiones_Q4.xlsx'))
                
            
        pbar.update(34)
        
        
        # #summary_com.rename(columns={0: "Importes_Q1"})
        # new_sum = pd.DataFrame({'Importes_Q4': list(summary_com[0])})
        # pbar.update(50)
        # summary_com.to_csv(os.path.join(ruta_cvul_res,'Comisiones_Q4.csv'))
        # fig1 = sns.barplot(data=new_sum.head(10), x=summary_com.index[:10], y=new_sum.columns[0])
        # plt.title('Comisiones 4º trimestre por producto', fontweight = 'bold')
        # plt.savefig(os.path.join(ruta_cvul_res,'G_Comisiones_Q4.jpg'))
        # plt.close()
        # pbar.update(50)
        
        
except Exception:
    print('NO SE HA PODIDO IMPORTAR EL ARCHIVO CVUL_Q4')













